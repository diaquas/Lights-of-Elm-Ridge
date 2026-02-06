#!/usr/bin/env python3
"""Export ModIQ matcher.ts scoring analysis to Excel workbook."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Styling helpers ──────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, num_cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT


def auto_width(ws, num_cols, max_width=60):
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    max_len = max(max_len, max(len(line) for line in lines))
        adjusted = min(max_len + 4, max_width)
        ws.column_dimensions[get_column_letter(col)].width = adjusted


def add_section_header(ws, row, text, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER
    return row + 1


# ─── Sheet 1: Base Factor Weights ──────────────────────────────

def create_base_weights(wb):
    ws = wb.create_sheet("Base Factor Weights")
    headers = ["Factor", "Weight", "Description", "Score Range"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Fuzzy Name Match", "40%", "Strongest signal — synonym expansion, token overlap, alias matching", "0.0 – 1.0"],
        ["Spatial Position", "25%", "Euclidean distance in normalized [0,1] layout space", "0.0 – 1.0"],
        ["Shape Classification", "15%", "Geometry: circular, linear, matrix, triangle, point, custom", "0.0 / 0.4 / 1.0"],
        ["Model Type (DisplayAs)", "12%", "xLights universal type from DisplayAs field", "0.0 / 0.3 / 0.7 / 1.0"],
        ["Node Count (Pixels)", "8%", "Pixel/node count similarity using drift tiers", "0.0 / 0.25 / 0.5 / 1.0"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_rows(ws, 2, len(data) + 1, len(headers))
    auto_width(ws, len(headers))


# ─── Sheet 2: Hard Exclusions ──────────────────────────────────

def create_hard_exclusions(wb):
    ws = wb.create_sheet("Hard Exclusions")
    headers = ["Rule", "Condition", "Result", "Exception"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["DMX Model", "type='DMX' or displayAs contains 'dmx' or name matches Pixel2DMX/Fog Machine/DMX Head", "Score = 0", "None"],
        ["Moving Head", "Name matches /\\bm\\.?h\\b/ or /moving\\s*head/", "Score = 0", "None"],
        ["Extreme Pixel Drift", "|srcPixels - destPixels| >= 1000 (non-groups)", "Score = 0", "Both are large spinners (500+ px each)"],
        ["Matrix Type-Lock", "One is Matrix, other is not", "Score = 0", "Both are groups"],
        ["Flood Type-Lock", "One is Flood but other is neither Flood nor Line", "Score = 0", "Both are groups"],
        ["Singing Type-Lock", "One is singing, other is not (detected via name/submodel patterns)", "Score = 0", "None"],
        ["Holiday Mismatch", "One is Halloween indicator, other is Christmas indicator", "Score = 0", "None"],
        ["GE Vendor Mismatch", "Both have GE prefix but different product lines", "Score = min(0.1, name*0.1)", "None — very low cap"],
        ["SUBMODEL_GROUP Lock", "One is SUBMODEL_GROUP, other is not", "Score = 0", "None"],
        ["Model vs Group Lock", "source.isGroup !== dest.isGroup", "Score = 0", "None"],
        ["Group Negation", "One group has 'no' (negation), other doesn't", "Score = 0", "None"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 3:
                cell.fill = RED_FILL

    style_data_rows(ws, 2, len(data) + 1, len(headers))
    auto_width(ws, len(headers))


# ─── Sheet 3: Name Scoring ────────────────────────────────────

def create_name_scoring(wb):
    ws = wb.create_sheet("Name Scoring (40%)")
    headers = ["Component", "Condition", "Score / Bonus", "Example"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Exact Normalized Match", "normalizeName(src) === normalizeName(dest)", "1.0", '"Arch 3" = "arch 3"'],
        ["Base Name Match", "baseName(src) === baseName(dest)", "0.85", '"Arch 1" ~ "ARCH 3" (both base="arch")'],
        ["Canonical Base Match", "canonicalBase(src) === canonicalBase(dest)", "0.85 (1.0 if same index)", '"Eave 1" ≡ "Horizontal 1" → 1.0'],
        ["Pixel Pole Synonym", "One is 'pole', other is 'pixel pole', same index", "1.0", '"Pole 3" ≡ "Pixel Pole 3"'],
        ["Singing Match", "Both isSinging(), same index", "1.0 (0.85 if diff index)", '"Singing Pumpkin 1" ≡ "Singing Skull 1"'],
        ["Token Overlap", "Bidirectional Jaccard with synonym expansion", "matches / max(|srcTokens|, |destTokens|)", '"Spinner Left" ~ "Spinner Right" = 0.67'],
        ["Substring Bonus", "srcBase contained in destNorm (or vice versa), len >= 3", "+0.15", '"showstopper" in "Showstopper Spinner Left"'],
        ["Prop Keyword Bonus", "Both names contain same PROP_KEYWORD", "+0.20", 'Both contain "arch"'],
        ["Alias Bonus", "source.aliases or dest.aliases match other's name", "+0.30", 'Old name via oldname: alias tag'],
        ["Semantic Synonym Boost", "When overlap < 0.5, calculateSynonymBoost() > 0.4", "+synScore × 0.25", '"Lawn Outline" ~ "Yard Border"'],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_rows(ws, 2, len(data) + 1, len(headers))

    # Add synonym table section
    row = len(data) + 3
    row = add_section_header(ws, row, "Synonym Map (SYNONYMS)", len(headers))

    synonyms = [
        ["l", "left, lft"],
        ["r", "right, rgt, rt"],
        ["c", "center, ctr, centre, mid, middle"],
        ["lg", "large, big"],
        ["sm", "small, mini, tiny"],
        ["grp", "group, all"],
        ["vert", "vertical, verticals, verts, verticle, verticl"],
        ["eave", "horizontal, horizontals, eaves, roofline, icicles"],
        ["arch", "arches, archway"],
        ["pole", "pixel pole"],
        ["cane", "candy cane, candycane, canes"],
        ["spin", "spinner"],
        ["moaw / boaw", "wreath"],
        ["mt", "mega tree, megatree"],
        ["mh", "moving head, movinghead"],
        ["ppd", "wreath"],
        ["tumba", "tombstone, tomb, tumbas"],
        ["estrella", "star, estrellas"],
        ["fantasma", "ghost, fantasmas"],
        ["calabaza", "pumpkin, calabazas"],
        ["fuzion", "rosa wreath, rosawreath, rosa"],
        ["chromaflake / icequeen", "snowflake, flake"],
        ["pimp", "singing pumpkin, singing face, singing prop"],
        ["singing pumpkin", "pimp, singing skull, singing face, singing prop"],
    ]
    syn_headers = ["Key", "Expands To"]
    for col, h in enumerate(syn_headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, 2)
    row += 1
    for syn in synonyms:
        for c, val in enumerate(syn, 1):
            ws.cell(row=row, column=c, value=val)
        row += 1
    style_data_rows(ws, row - len(synonyms), row - 1, 2)

    # Equivalent bases
    row += 1
    row = add_section_header(ws, row, "Equivalent Bases (EQUIVALENT_BASES)", len(headers))
    equiv = [
        ["eave / eaves", "structural_horizontal"],
        ["horizontal / horizontals", "structural_horizontal"],
        ["vert / verts / verticle / verticl", "structural_vertical"],
        ["vertical / verticals", "structural_vertical"],
    ]
    eq_headers = ["Variants", "Canonical Form"]
    for col, h in enumerate(eq_headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, 2)
    row += 1
    for eq in equiv:
        for c, val in enumerate(eq, 1):
            ws.cell(row=row, column=c, value=val)
        row += 1
    style_data_rows(ws, row - len(equiv), row - 1, 2)

    auto_width(ws, len(headers))


# ─── Sheet 4: Spatial Scoring ─────────────────────────────────

def create_spatial_scoring(wb):
    ws = wb.create_sheet("Spatial Scoring (25%)")
    headers = ["Condition", "Score", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Groups", "0.5 (neutral)", "Groups don't have meaningful spatial positions"],
        ["Euclidean distance = 0", "1.0", "Identical position in normalized layout space"],
        ["Euclidean distance = √2 (max)", "0.0", "Opposite corners of the layout"],
        ["Formula", "1.0 - dist / 1.414", "Linear interpolation between min and max distance"],
        ["Normalization", "Per-axis [0,1]", "Each axis independently normalized to layout bounds: (value - min) / (max - min)"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_rows(ws, 2, len(data) + 1, len(headers))
    auto_width(ws, len(headers))


# ─── Sheet 5: Shape Scoring ──────────────────────────────────

def create_shape_scoring(wb):
    ws = wb.create_sheet("Shape Scoring (15%)")
    headers = ["Shape", "Detection Rules (DisplayAs / Name patterns)", "Score if Match", "Score if Related", "Score Otherwise"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["circular", "DA: circle/spinner/sphere/wreaths; Name: spinner|wreath|circle|ring|globe|ball|rosa|fuzion|overlord|starburst", "1.0", "0.4 (→custom)", "0.0"],
        ["matrix", "DA: *matrix*/cube/window frame; Name: matrix|panel|p5|p10|fence|window|sign|tune.*to", "1.0", "0.4 (→custom)", "0.0"],
        ["linear", "DA: single line/poly line/icicles/arches/candy cane(s); Name: eave|vert|horizontal|roofline|outline|driveway|pole|cane|icicle|arch", "1.0", "0.4 (→custom)", "0.0"],
        ["triangle", "DA: *tree*; Name: tree|mega.*tree|spiral|firework", "1.0", "0.4 (→custom)", "0.0"],
        ["point", "DA: star; Name: star|flood|bulb", "1.0", "0.4 (→custom)", "0.0"],
        ["custom", "Fallback when no other shape matches", "1.0", "0.4 (→any)", "0.0"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_rows(ws, 2, len(data) + 1, len(headers))
    auto_width(ws, len(headers))


# ─── Sheet 6: Type Scoring ───────────────────────────────────

def create_type_scoring(wb):
    ws = wb.create_sheet("Type Scoring (12%)")
    headers = ["Condition", "Score"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Exact type match (source.type === dest.type)", "1.0"],
        ["Both are groups (isGroup)", "0.7"],
        ["Related type (RELATED_TYPES map)", "0.7"],
        ["One is 'Custom' type", "0.3"],
        ["No relationship", "0.0"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_rows(ws, 2, len(data) + 1, len(headers))

    # Related types table
    row = len(data) + 3
    row = add_section_header(ws, row, "RELATED_TYPES Map", 2)
    related = [
        ["Tree", "Mega Tree, Spiral Tree"],
        ["Mega Tree", "Tree"],
        ["Spiral Tree", "Tree"],
        ["Arch", "Candy Cane"],
        ["Candy Cane", "Arch"],
        ["Spinner", "Wreath, Snowflake"],
        ["Wreath", "Spinner"],
        ["Spider / Bat / Tombstone", "Custom"],
        ["Pumpkin", "Custom, Ghost"],
        ["Ghost", "Custom, Pumpkin"],
        ["Matrix / Fence / Sign / Window", "(none)"],
        ["Line", "Roofline, Outline, Flood"],
        ["Roofline", "Line"],
        ["Outline", "Line, Roofline"],
        ["Pole", "Line"],
        ["Flood", "Line"],
        ["Pixel Forest", "Matrix, Fence"],
        ["Singing Face", "Custom"],
        ["Star", "Snowflake"],
        ["Snowflake", "Star, Spinner"],
    ]
    rel_headers = ["Type", "Related Types (score 0.7)"]
    for col, h in enumerate(rel_headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, 2)
    row += 1
    for rel in related:
        for c, val in enumerate(rel, 1):
            ws.cell(row=row, column=c, value=val)
        row += 1
    style_data_rows(ws, row - len(related), row - 1, 2)
    auto_width(ws, 2)


# ─── Sheet 7: Pixel Scoring ──────────────────────────────────

def create_pixel_scoring(wb):
    ws = wb.create_sheet("Pixel Scoring (8%)")
    headers = ["Drift (|src - dest|)", "Score", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["0", "1.0", "Exact pixel count match"],
        ["1 – 99", "0.5", "Minor variation (different strand config)"],
        ["100 – 499", "0.25", "Moderate difference"],
        ["≥ 500", "0.0", "Fundamentally different props"],
        ["Groups", "0.5 (neutral)", "Groups don't have meaningful pixel counts"],
        ["Either = 0", "0.5 (neutral)", "Unknown pixel count"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_rows(ws, 2, len(data) + 1, len(headers))
    auto_width(ws, len(headers))


# ─── Sheet 8: Score Overrides & Modifiers ─────────────────────

def create_overrides(wb):
    ws = wb.create_sheet("Score Overrides & Modifiers")
    headers = ["Override / Modifier", "Condition", "Effect", "Priority"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Pole/Pixel Pole Force", "isPolePair() — base names match after stripping 'pixel', same index", "Force score=1.0, bypass all hard exclusions", "Highest — before any exclusion"],
        ["GE Product Mismatch Cap", "Both have GE prefix, different product lines", "Score capped at min(0.1, name×0.1)", "After hard exclusions"],
        ["SUBMODEL_GROUP Semantic", "Both SUBMODEL_GROUP, same semanticCategory", "Score = 0.75 + name×0.25", "Group matching"],
        ["SUBMODEL_GROUP Category Penalty", "Both SUBMODEL_GROUP, different semanticCategory", "Score = name × 0.5", "Group matching"],
        ["Group Scoring", "Both isGroup (non-submodel)", "Score = (name×0.55 + memberOverlap×0.30 + type×0.15) × 1.3", "Group matching"],
        ["House-Location Reweighting", "Either model has house location term (office, garage, etc.)", "Name: 25%, Spatial: 25%, Shape: 15%, Type: 27%, Pixels: 8%", "Individual models only"],
        ["Coordinate Tiebreaker", "Same base name, different index (e.g., 'Window 1' vs 'Window 2')", "Name: 20%, Spatial: 45%, Shape: 15%, Type: 12%, Pixels: 8%", "Individual models only"],
        ["Eave/Vert Penalty", "Eave or vertical model matched to non-house-line model", "Score × 0.4", "Post-weighting"],
        ["Vendor Pixel Hint", "Both share exact pixel count matching known vendor product", "Score × 1.15 (capped at 1.0)", "Post-weighting"],
        ["Large Spinner Exception", "Both isLargeSpinner() and >= 500px each", "Exempt from 1000+ pixel drift exclusion", "Hard exclusion exception"],
        ["Spinner Shared-Source", "Both spinners with ≥3 matching submodel names", "Source not consumed (allows 1-to-many mapping)", "Greedy assignment"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_rows(ws, 2, len(data) + 1, len(headers))

    # Vendor pixel hints
    row = len(data) + 3
    row = add_section_header(ws, row, "Vendor Pixel Fingerprints (VENDOR_PIXEL_HINTS)", len(headers))
    hints = [
        ["269", "ccc_spinner_18", "CCC Spinner 18\""],
        ["451", "ccc_spinner_24", "CCC Spinner 24\""],
        ["519", "ccc_spinner_25", "CCC Spinner 25\""],
        ["596", "ccc_spinner_36", "CCC Spinner 36\""],
        ["640", "ge_flake_640", "GE Flake 640px"],
        ["800", "efl_showstopper", "EFL Showstopper"],
        ["1046", "ccc_spinner_48", "CCC Spinner 48\""],
        ["1117", "boscoyo_mesmerizer", "Boscoyo Mesmerizer"],
        ["1529", "ge_overlord", "GE Overlord"],
    ]
    hint_headers = ["Pixel Count", "Product ID", "Description"]
    for col, h in enumerate(hint_headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, 3)
    row += 1
    for hint in hints:
        for c, val in enumerate(hint, 1):
            ws.cell(row=row, column=c, value=val)
        row += 1
    style_data_rows(ws, row - len(hints), row - 1, 3)
    auto_width(ws, len(headers))


# ─── Sheet 9: Phase 3 Cross-Prop Matching ────────────────────

def create_phase3(wb):
    ws = wb.create_sheet("Phase 3 Cross-Prop")
    headers = ["Class", "Member Keywords", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["halloween_yard", "pumpkin, ghost, tombstone, skeleton, spider, bat, skull, witch, zombie, reaper, scarecrow, cauldron, coffin, black cat", "Cross-matches unmapped Halloween props"],
        ["christmas_yard", "snowman, present, gift, candy cane, stocking, nutcracker, ornament", "Cross-matches unmapped Christmas props"],
        ["mini_yard", "mini tree, ghost, black cat, mini pumpkin", "Cross-holiday mini props"],
        ["tree", "tree, mega tree, megatree, spiral", "Tree variants"],
        ["arch", "arch, archway, candy cane, cane", "Arch/cane variants"],
        ["star_flake", "star, snowflake, flake", "Star and snowflake interop"],
        ["wreath_spinner", "wreath, spinner, rosa, fuzion", "Wreath/spinner interop"],
        ["structural_line", "eave, vertical, roofline, outline, horizontal", "Structural line elements"],
        ["vertical_structure", "pole, vertical, fence", "Vertical structures"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_rows(ws, 2, len(data) + 1, len(headers))

    row = len(data) + 3
    row = add_section_header(ws, row, "Phase 3 Scoring Rules", len(headers))
    rules = [
        ["Individual Models (Phase 3)", "Score = (name×0.6 + pixels×0.4) × 0.7", "Pre-penalty threshold: > 0.57; must reach LOW (0.40) after penalty"],
        ["Groups (Phase 3b)", "Score = (name×0.4 + memberOverlap×0.4 + type×0.2) × 0.7", "Same threshold; SUBMODEL_GROUP compatibility enforced"],
        ["Cross-Prop Penalty", "All Phase 3 scores × 0.7", "Caps cross-prop matches below HIGH confidence"],
    ]
    rule_headers = ["Rule", "Formula", "Notes"]
    for col, h in enumerate(rule_headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, 3)
    row += 1
    for rule in rules:
        for c, val in enumerate(rule, 1):
            ws.cell(row=row, column=c, value=val)
        row += 1
    style_data_rows(ws, row - len(rules), row - 1, 3)
    auto_width(ws, len(headers))


# ─── Sheet 10: Confidence Thresholds ─────────────────────────

def create_confidence(wb):
    ws = wb.create_sheet("Confidence Thresholds")
    headers = ["Tier", "Score Range", "Color", "Description"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["High", "≥ 0.85", "Green", "Very confident match — likely correct"],
        ["Medium", "0.60 – 0.84", "Yellow", "Reasonable match — should verify"],
        ["Low", "0.40 – 0.59", "Orange", "Weak match — review carefully"],
        ["Unmapped", "< 0.40", "Red", "No suitable match found"],
    ]
    fills = [GREEN_FILL, YELLOW_FILL, ORANGE_FILL, RED_FILL]
    for r, (row_data, fill) in enumerate(zip(data, fills), 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fill

    style_data_rows(ws, 2, len(data) + 1, len(headers))

    row = len(data) + 3
    row = add_section_header(ws, row, "Auto-Accept Phase Routing (V4 Wizard)", len(headers))
    routing = [
        ["Auto-Accept Phase", "≥ 0.70", "—", "All entity types (models, groups, spinners) with 70%+ go to auto-accept for bulk opt-out review"],
        ["Groups Phase", "< 0.70", "—", "MODEL_GROUP + META_GROUP + MIXED_GROUP below threshold"],
        ["Models Phase", "< 0.70", "—", "MODEL + SUBMODEL below threshold"],
        ["Spinners Phase", "< 0.70", "—", "SUBMODEL_GROUP below threshold"],
        ["Review Phase", "All", "—", "Shows all items regardless of score"],
    ]
    for row_data in routing:
        for c, val in enumerate(row_data, 1):
            ws.cell(row=row, column=c, value=val)
        row += 1
    style_data_rows(ws, row - len(routing), row - 1, len(headers))
    auto_width(ws, len(headers))


# ─── Sheet 11: Holiday Detection ─────────────────────────────

def create_holiday(wb):
    ws = wb.create_sheet("Holiday Detection")
    headers = ["Holiday", "Indicator Keywords"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Halloween", "spider, bat, ghost, pumpkin, tombstone, tomb, skull, skeleton, witch, zombie, reaper, scarecrow, cauldron, coffin, graveyard, rip"],
        ["Christmas", "wreath, snowflake, flake, bow, candy cane, cane, snowman, stocking, nutcracker, ornament, present, gift, sleigh, reindeer, santa, angel, nativity, grinch, elf"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_rows(ws, 2, len(data) + 1, len(headers))
    auto_width(ws, len(headers))


# ─── Sheet 12: Matching Algorithm ────────────────────────────

def create_algorithm(wb):
    ws = wb.create_sheet("Matching Algorithm")
    headers = ["Phase", "Source Pool", "Dest Pool", "Scoring", "Post-Processing"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Phase 1: Groups", "sourceModels.filter(isGroup)", "destModels.filter(isGroup)", "computeScore() with group-specific weighting", "Group-only children detection: lower confidence if no exposed members"],
        ["Phase 2: Individuals", "sourceModels.filter(!isGroup)", "Remaining dest (unmatched groups + individuals)", "computeScore() with standard weights", "—"],
        ["Phase 3: Cross-Prop Individuals", "Unmapped non-group sources", "Unused non-group dests", "name×0.6 + pixels×0.4, then ×0.7 penalty", "Must reach LOW (0.40) after penalty"],
        ["Phase 3b: Cross-Prop Groups", "Unmapped group sources", "Unused group dests", "name×0.4 + members×0.4 + type×0.2, then ×0.7", "SUBMODEL_GROUP compatibility enforced"],
        ["Greedy Assignment", "All phases", "All phases", "Score matrix sorted descending, greedy 1:1", "Spinner shared-source: 1:many when submodels match"],
        ["Minimum Score", "—", "—", "> 0.1 to enter score matrix", "confidence='unmapped' assignments skipped"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_rows(ws, 2, len(data) + 1, len(headers))

    row = len(data) + 3
    row = add_section_header(ws, row, "Output Sorting", len(headers))
    sort_data = [
        ["Primary", "Confidence tier: high → medium → low → unmapped"],
        ["Secondary", "Groups before individuals within same tier"],
        ["Tertiary", "Natural number sort on source model name (e.g., 'Eave 2' < 'Eave 10')"],
    ]
    for sort_row in sort_data:
        ws.cell(row=row, column=1, value=sort_row[0])
        ws.cell(row=row, column=2, value=sort_row[1])
        row += 1

    auto_width(ws, len(headers))


# ─── Main ─────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    create_base_weights(wb)
    create_hard_exclusions(wb)
    create_name_scoring(wb)
    create_spatial_scoring(wb)
    create_shape_scoring(wb)
    create_type_scoring(wb)
    create_pixel_scoring(wb)
    create_overrides(wb)
    create_phase3(wb)
    create_confidence(wb)
    create_holiday(wb)
    create_algorithm(wb)

    output = "/home/user/Lights-of-Elm-Ridge/ModIQ_Matcher_Analysis.xlsx"
    wb.save(output)
    print(f"Saved: {output}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
