#!/usr/bin/env python3
"""
Generate ModIQ V5.2 Matching Rules Reference Excel file.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Style definitions ──────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
BODY_FONT = Font(name="Calibri", size=10)
MONO_FONT = Font(name="Consolas", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LOW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def style_body(ws, num_rows, num_cols, mono_cols=None):
    mono_cols = mono_cols or []
    for row in range(2, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = MONO_FONT if col in mono_cols else BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER


def auto_width(ws, num_cols, max_width=60):
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    max_len = max(max_len, max(len(l) for l in lines))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, max_width)


# ═══════════════════════════════════════════════════════════════
# SHEET 1: V5.2 Enhancements (the 10 tickets)
# ═══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "V5.2 Enhancements"
headers = ["Ticket #", "Rule Name", "Impact", "Effort", "Est. Match Gain",
           "File Changed", "Description", "Examples (Before → After)", "Revert Command"]
for c, h in enumerate(headers, 1):
    ws.cell(row=1, column=c, value=h)

tickets = [
    [1, "Plural/Singular Normalization", "HIGH", "Easy", "+3-5%",
     "matcher.ts",
     "singularize() function added to normalizeName(). Automatically reduces plurals to singular form so 'Arches' matches 'Arch' without needing explicit synonym entries.",
     "Arches→arch, Tombstones→tombstone, Spiders→spider, Bushes→bush, Witches→witch, Circles→circle",
     "git revert 81d7c17"],
    [2, "Compound Word Splitting", "HIGH", "Easy", "+2-4%",
     "matcher.ts",
     "splitCompoundWords() splits PascalCase/camelCase names before lowercasing. Preserves uppercase acronyms like DMX, GE.",
     "MegaTree→mega tree, CandyCane→candy cane, RoofLine→roof line, PixelPole→pixel pole, DMXHead→DMX Head",
     "git revert e08d6be"],
    [3, "Expanded EQUIVALENT_BASES", "HIGH", "Easy", "+2-3%",
     "matcher.ts",
     "Added 30+ canonical base-name equivalences. Words with different spellings for the same concept now auto-canonicalize.",
     "arch=archway=arc, flood=wash, spinner=pinwheel, wreath=rosa, snowflake=flake=chromaflake, driveway=sidewalk=walkway",
     "git revert 194400b"],
    [4, "Enhanced Name Prefix Stripping", "HIGH", "Easy", "+2-4%",
     "matcher.ts",
     "Strips channel/universe/port/bracket prefixes common in community xLights layouts. 7 new regex patterns.",
     "CH01 Arch 1→Arch 1, U1-Arch→Arch, P1S1 - Arch→Arch, [1] Arch→Arch, (01) Arch→Arch, 01 - Arch→Arch",
     "git revert 3fee4ad"],
    [5, "Relaxed Pixel Drift (Same-Type)", "MEDIUM", "Easy", "+1-3%",
     "matcher.ts",
     "When name score ≥0.85 AND type score ≥0.7, pixel factor is floored at 0.5. Prevents pixel count differences from dragging obvious matches into lower tiers.",
     "50px arch + 100px arch: before=low conf, after=high conf. Same prop, different pixel density.",
     "git revert 4141483"],
    [6, "Surplus-to-Spatial Matching", "HIGH", "Medium", "+3-6%",
     "matcher.ts",
     "New Phase 3c: surplus unmapped source models → spatially nearest already-matched dest of same base name. Many-to-one mapping so effects don't disappear when prop counts differ.",
     "Source: 8 arches, User: 5 arches. Before: arches 6-8 unmapped. After: each maps to nearest user arch (medium confidence).",
     "git revert 1a76796"],
    [7, "DisplayAs TYPE_MAP Gaps", "MEDIUM", "Easy", "+1-2%",
     "parser.ts",
     "Added missing xLights DisplayAs values (Tree 180/270, Horiz Matrix, Sphere, Cube, DMX Flood/Servo) and 15+ new inferTypeFromName patterns. Reordered by specificity.",
     "Tree 180→Tree, Horiz Matrix→Matrix, Sphere→Circle, snowman→Custom, deer→Custom, net/mesh→Matrix, stake→Pole",
     "git revert cb80155"],
    [8, "Bidirectional Related Types", "MEDIUM", "Easy", "+1-2%",
     "matcher.ts",
     "Expanded RELATED_TYPES map and Shape cross-compatibility. More type pairs now score 0.7 (related) instead of 0.0.",
     "Circle↔Wreath↔Spinner, Icicles↔Line↔Roofline, Matrix↔Fence, Star↔Circle, Snowflake↔Wreath. Shapes: circular↔point, linear↔triangle.",
     "git revert d746bd3"],
    [9, "Abbreviation Chain Resolution", "MEDIUM", "Medium", "+1-2%",
     "matcher.ts",
     "Two-pass synonym expansion in tokenize(). Pass 1 expands direct synonyms, Pass 2 expands new tokens found in Pass 1. Catches multi-level abbreviation chains.",
     "SS L 1 → (pass1) showstopper left 1 → (pass2) show stopper left 1. Multi-hop: mt → mega tree → megatree.",
     "git revert 65b3bb2"],
    [10, "More Vendor Pixel Hints", "MEDIUM", "Easy", "+1-2%",
     "matcher.ts",
     "Added 20+ new entries to VENDOR_PIXEL_HINTS covering GE, EFL, Boscoyo, Holiday Coro products and standard P5/P10 panel sizes.",
     "GE Rosa Grande=1200px, GE Fuzion=960px, EFL BabyFlake=400px, P10 32x32=1024px, P5 64x64=4096px",
     "git revert 584c646"],
]

for r, row in enumerate(tickets, 2):
    for c, val in enumerate(row, 1):
        ws.cell(row=r, column=c, value=val)
    # Color-code impact
    impact_cell = ws.cell(row=r, column=3)
    if row[2] == "HIGH":
        impact_cell.fill = HIGH_FILL
    elif row[2] == "MEDIUM":
        impact_cell.fill = MED_FILL

style_header(ws, len(headers))
style_body(ws, len(tickets) + 1, len(headers), mono_cols=[9])
auto_width(ws, len(headers))

# ═══════════════════════════════════════════════════════════════
# SHEET 2: Scoring Factors (the 6 weighted factors)
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Scoring Factors")
headers2 = ["Factor", "Weight", "Weight %", "Description", "Score Range", "Key Logic", "Examples"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)

factors = [
    ["Name", 0.38, "38%",
     "Fuzzy name matching — strongest signal. Combines exact match, base-name match, token overlap with synonym expansion, substring containment, prop keyword match, alias bonus, and semantic synonym boost.",
     "0.0 – 1.0",
     "Exact normalized = 1.0, Base name match = 0.85, Canonical base + same index = 1.0, Token overlap = proportional, Alias bonus = +0.30, Semantic synonym boost = +synScore*0.25",
     "\"Arch 1\" vs \"Arch 1\" = 1.0\n\"Eave 1\" vs \"Horizontal 1\" = 1.0 (canonical)\n\"Lawn Outline\" vs \"Yard Border\" = high (synonym boost)"],
    ["Spatial", 0.22, "22%",
     "Euclidean distance in normalized [0,1] coordinate space. Primary disambiguator for identically-named props (Arch 1 vs Arch 2). Reweighted to 42% when both models share a base name with different indices.",
     "0.0 – 1.0",
     "score = 1.0 - distance (clamped to [0,1]). Distance uses WorldPosX/Y normalized to layout bounds.",
     "Props at same position = 1.0\nProps across yard = ~0.3\nReweighted to 42% for index disambiguation"],
    ["Shape", 0.13, "13%",
     "Geometry classification into 6 classes: circular, linear, matrix, triangle, point, custom. Cross-compatibility between related shapes.",
     "0.0 / 0.4 / 1.0",
     "Same shape = 1.0, Related shapes (e.g., circular↔point) = 0.4, Unrelated = 0.0. Groups = 0.5 (neutral).",
     "Spinner vs Wreath = 1.0 (both circular)\nArch vs Tree = 0.4 (linear↔triangle)\nSpinner vs Matrix = 0.0"],
    ["Type", 0.10, "10%",
     "xLights DisplayAs type comparison. Uses RELATED_TYPES table for partial credit.",
     "0.0 / 0.3 / 0.7 / 1.0",
     "Same type = 1.0, Related type = 0.7, One/both Custom = 0.3, Unrelated = 0.0. Both groups = 0.7.",
     "Arch vs Arch = 1.0\nArch vs Candy Cane = 0.7 (related)\nSpinner vs Matrix = 0.0"],
    ["Pixels", 0.10, "10%",
     "Node/pixel count similarity via min/max ratio curve. Penalizes large differences but floored at 0.5 when name+type match well (Ticket 5).",
     "0.0 – 1.0",
     "ratio ≥0.95 → ~1.0, ≥0.80 → 0.60-0.90, ≥0.50 → 0.20-0.60, ≥0.25 → 0.0-0.20, <0.25 → 0.0. Groups/zero px = 0.5.",
     "100px vs 105px = ~0.995\n100px vs 200px = 0.20\n50px arch vs 100px arch = 0.50 (floored, Ticket 5)"],
    ["Structure", 0.07, "7%",
     "Submodel count similarity. Penalizes models where one has submodels and the other doesn't.",
     "0.2 – 1.0",
     "Neither has subs = 0.5, One has/other doesn't = 0.2, Both have subs = ratio-based (0.2-1.0). Groups = 0.5.",
     "12 subs vs 12 subs = 1.0\n12 subs vs 8 subs = 0.5\nSpider (12 subs) vs Arch (0 subs) = 0.2"],
]

for r, row in enumerate(factors, 2):
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)

style_header(ws2, len(headers2))
style_body(ws2, len(factors) + 1, len(headers2))
auto_width(ws2, len(headers2))

# ═══════════════════════════════════════════════════════════════
# SHEET 3: Hard Exclusions
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Hard Exclusions")
headers3 = ["#", "Rule Name", "Condition", "Result", "Bypass", "Example Blocked"]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=c, value=h)

exclusions = [
    [1, "DMX Exclusion", "type='DMX', displayAs contains 'dmx', or name matches pixel2dmx/dmx head/fog machine", "Score=0, all factors zeroed", "None", "Pixel2DMX, Fog Machine, DmxGeneral"],
    [2, "Moving Head Exclusion", "Name matches /\\bm\\.?h\\b|moving\\s*head/i", "Score=0, all factors zeroed", "None", "M.H. Spot 1, Moving Head 3"],
    [3, "Extreme Pixel Drift", "Both non-group, both pixelCount>0, abs diff ≥1000", "Score=0", "Both are large spinners (≥500px each)", "100px spider vs 12000px matrix → blocked\n800px Showstopper vs 1529px Overlord → ALLOWED"],
    [4, "Matrix Type-Lock", "One is matrix-type, other is not", "Score=0", "Both are groups", "Matrix cannot match Arch. Matrix group CAN match non-matrix group."],
    [5, "Flood Type-Lock", "One/both is flood; must be both-flood or flood+line", "Score=0", "Both are groups", "Flood cannot match Spinner. Flood CAN match Line."],
    [6, "Singing Mismatch", "One is singing (pimp/efl snowman/mouth submodel), other is not", "Score=0, all factors zeroed", "None", "Singing Pumpkin blocked from matching Arch 1"],
    [7, "Holiday Mismatch", "Both have holiday indicators but different holidays", "Score=0", "Only one model has indicators", "Halloween Ghost vs Christmas Snowman"],
    [8, "GE Product Mismatch", "Both have GE prefix but different product lines", "Score capped at min(0.1, name*0.1)", "None", "GE Fuzion 1 vs GE Rosa Grande 2"],
    [9, "SUBMODEL_GROUP Lock", "One is SUBMODEL_GROUP and other is not", "Score=0, all factors zeroed", "None", "Spinner submodel group vs regular MODEL_GROUP"],
    [10, "Model vs Group", "source.isGroup !== dest.isGroup", "Score=0, all factors zeroed", "None", "Individual Spiral 2 vs 09 All Mega Trees group"],
    [11, "Group Negation (NO)", "One group has 'no' negation, other doesn't", "Score=0", "None", "All - No Spinners vs All Spinners"],
    [12, "Pole/Pixel Pole Bypass", "Both are pole-type, one has 'pixel' prefix, indices match", "FORCE score=1.0, all factors maxed", "N/A (this IS a bypass)", "Pole 3 ↔ Pixel Pole 3 ALWAYS match"],
]

for r, row in enumerate(exclusions, 2):
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)

style_header(ws3, len(headers3))
style_body(ws3, len(exclusions) + 1, len(headers3))
auto_width(ws3, len(headers3))

# ═══════════════════════════════════════════════════════════════
# SHEET 4: Matching Phases
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Matching Phases")
headers4 = ["Phase", "Name", "Scope", "Algorithm", "Scoring Formula", "Threshold", "Confidence Tiers"]
for c, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=c, value=h)

phases = [
    ["Phase 1", "Group Matching", "Source groups → Dest groups",
     "Greedy assignment by descending score. Groups matched first to establish coverage.",
     "name*0.55 + memberOverlap*0.30 + type*0.15, then ×1.3 boost (capped at 1.0)",
     ">0.1 to enter matrix", "≥0.85=high, ≥0.60=medium, ≥0.40=low, <0.40=unmapped"],
    ["Phase 2", "Individual Matching", "Source individuals → Remaining dest individuals",
     "Greedy assignment by descending score. Spinner reuse: matched spinners with ≥3 shared submodel names remain available for additional dest spinners.",
     "name*0.38 + spatial*0.22 + shape*0.13 + type*0.10 + pixels*0.10 + structure*0.07",
     ">0.1 to enter matrix", "≥0.85=high, ≥0.60=medium, ≥0.40=low, <0.40=unmapped"],
    ["Post", "Group Children Check", "Matched dest groups with members",
     "If a matched dest group's member names don't exist as individual models in dest layout, confidence lowered by one tier.",
     "N/A (modifier only)", "N/A",
     "high→medium, medium→low. Adds reason: 'Group children not exposed in layout'"],
    ["Phase 3", "Quantity Matching (Individuals)", "Unmapped source individuals → Unused dest individuals",
     "Cross-match within same INTERCHANGEABLE_CLASS (e.g., halloween_yard, tree, arch). Ordinal matching by index.",
     "(name*0.6 + pixels*0.4) × 0.7 penalty",
     "Pre-penalty >0.57", "Same tiers. Reason: 'Quantity match (class name)'"],
    ["Phase 3c", "Surplus-to-Spatial (NEW V5.2)", "Still-unmapped source individuals → Already-matched dest models",
     "Map surplus source models to spatially nearest already-matched dest of same base name or canonical base. Many-to-one mapping.",
     "Fixed score = 0.65 (medium confidence). Full factor breakdown computed for UI.",
     "Must share base name or canonical base", "Always medium. Reason: 'Surplus → nearest same-type prop'"],
    ["Phase 3b", "Group Interchangeability", "Unmapped source groups → Unused dest groups",
     "Cross-match within same INTERCHANGEABLE_CLASS. SUBMODEL_GROUP must match SUBMODEL_GROUP only.",
     "(name*0.4 + memberOverlap*0.4 + type*0.2) × 0.7 penalty",
     "Pre-penalty >0.57", "Same tiers. Reason: 'Quantity match (class name)'"],
]

for r, row in enumerate(phases, 2):
    for c, val in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=val)

style_header(ws4, len(headers4))
style_body(ws4, len(phases) + 1, len(headers4))
auto_width(ws4, len(headers4))

# ═══════════════════════════════════════════════════════════════
# SHEET 5: Synonyms & Abbreviations
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Synonyms & Abbreviations")
headers5 = ["Key (Abbreviation)", "Expands To", "Category", "Example Match"]
for c, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=c, value=h)

synonyms = [
    ["l", "left, lft", "Direction", "L Arch → Left Arch"],
    ["r", "right, rgt, rt", "Direction", "R Arch → Right Arch"],
    ["c", "center, ctr, centre, mid, middle", "Direction", "C Arch → Center Arch"],
    ["lt", "left", "Direction", "Lt Tree → Left Tree"],
    ["lg", "large, big", "Size", "Lg Star → Large Star"],
    ["sm", "small, mini, tiny", "Size", "Sm Tree → Small Tree"],
    ["med", "medium", "Size", "Med Arch → Medium Arch"],
    ["ss", "showstopper", "Vendor", "SS 1 → Showstopper 1"],
    ["mt", "mega tree, megatree", "Compound", "MT 1 → Mega Tree 1"],
    ["mh", "moving head, movinghead", "Compound", "MH 1 → Moving Head 1"],
    ["dw", "driveway", "Location", "DW → Driveway"],
    ["sw", "sidewalk", "Location", "SW → Sidewalk"],
    ["ppd", "wreath", "Vendor", "PPD → Wreath"],
    ["moaw", "wreath", "Vendor", "MOAW → Wreath"],
    ["boaw", "wreath", "Vendor", "BOAW → Wreath"],
    ["spin", "spinner", "Type", "Spin 1 → Spinner 1"],
    ["vert", "vertical, verticals, verts, verticle, verticl", "Structural", "Vert 1 → Vertical 1"],
    ["horiz", "horizontal, horizontals", "Structural", "Horiz 1 → Horizontal 1"],
    ["eave", "horizontal, horizontals, eaves, roofline, icicles", "Structural", "Eave 1 → Roofline 1"],
    ["roofline", "eave, eaves, horizontal, horizontals", "Structural", "Roofline 1 → Eave 1"],
    ["arch", "arches, archway", "Props", "Arch → Arches / Archway"],
    ["cane", "candy cane, candycane, canes", "Props", "Cane 1 → Candy Cane 1"],
    ["pole", "pixel pole", "Props", "Pole 1 → Pixel Pole 1"],
    ["pimp", "singing pumpkin, singing face, singing prop", "Singing", "Pimp 1 → Singing Pumpkin 1"],
    ["tumba", "tombstone, tomb, tumbas", "Spanish", "Tumba 1 → Tombstone 1"],
    ["estrella", "star, estrellas", "Spanish", "Estrella 1 → Star 1"],
    ["arbol", "tree, arboles", "Spanish", "Arbol 1 → Tree 1"],
    ["contorno", "outline, contornos", "Spanish", "Contorno → Outline"],
    ["calabaza", "pumpkin, calabazas", "Spanish", "Calabaza 1 → Pumpkin 1"],
    ["fantasma", "ghost, fantasmas", "Spanish", "Fantasma 1 → Ghost 1"],
    ["araña", "spider, arañas", "Spanish", "Araña 1 → Spider 1"],
    ["firework", "spiral tree, inverted tree, spiral, fireworks", "Cross-Holiday", "Firework 1 → Spiral Tree 1"],
    ["pixel forest", "peace stakes, vertical matrix, icicles", "Cross-Holiday", "Pixel Forest → Peace Stakes"],
    ["mini trees", "ghosts, black cats", "Halloween/Xmas Swap", "Mini Trees → Ghosts"],
    ["fuzion", "rosa wreath, rosawreath, rosa", "GE Product", "GE Fuzion → Rosa Wreath"],
    ["chromaflake", "snowflake, flake", "Vendor", "ChromaFlake → Snowflake"],
    ["chromatrim", "outline, roofline, trim", "Vendor", "ChromaTrim → Outline"],
    ["fence", "vertical matrix", "Structural", "Fence → Vertical Matrix"],
]

for r, row in enumerate(synonyms, 2):
    for c, val in enumerate(row, 1):
        ws5.cell(row=r, column=c, value=val)

style_header(ws5, len(headers5))
style_body(ws5, len(synonyms) + 1, len(headers5))
auto_width(ws5, len(headers5))

# ═══════════════════════════════════════════════════════════════
# SHEET 6: EQUIVALENT_BASES
# ═══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Equivalent Bases")
headers6 = ["Variant Names", "Canonical Form", "Category", "Example"]
for c, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=c, value=h)

equiv = [
    ["eave, eaves, horizontal, horizontals, roofline, gutter", "structural_horizontal", "Horizontal structural", "\"Eave 1\" = \"Roofline 1\" = \"Horizontal 1\""],
    ["vert, verts, vertical, verticals, verticle, verticl", "structural_vertical", "Vertical structural", "\"Vert 1\" = \"Vertical 1\" = \"Verticle 1\""],
    ["outline, border, trim, perimeter", "structural_outline", "Outline/border", "\"Outline\" = \"Border\" = \"Trim\""],
    ["arch, archway, arc", "prop_arch", "Arch props", "\"Arch 1\" = \"Archway 1\""],
    ["cane, candycane, candy cane", "prop_cane", "Candy cane", "\"Cane 1\" = \"CandyCane 1\""],
    ["megatree, mega tree", "prop_megatree", "Mega tree", "\"MegaTree\" = \"Mega Tree\""],
    ["pole, pixel pole", "prop_pole", "Pole", "\"Pole 1\" = \"Pixel Pole 1\""],
    ["spinner, pinwheel", "prop_spinner", "Spinner", "\"Spinner 1\" = \"Pinwheel 1\""],
    ["wreath, rosa", "prop_wreath", "Wreath", "\"Wreath 1\" = \"Rosa 1\""],
    ["flood, wash", "prop_flood", "Flood/wash", "\"Flood 1\" = \"Wash 1\""],
    ["stake, rod", "prop_stake", "Stake", "\"Stake 1\" = \"Rod 1\""],
    ["snowflake, flake, chromaflake", "prop_snowflake", "Snowflake", "\"Snowflake\" = \"Flake\" = \"ChromaFlake\""],
    ["driveway, sidewalk, walkway, pathway", "structural_pathway", "Pathway", "\"Driveway\" = \"Sidewalk\" = \"Walkway\""],
]

for r, row in enumerate(equiv, 2):
    for c, val in enumerate(row, 1):
        ws6.cell(row=r, column=c, value=val)

style_header(ws6, len(headers6))
style_body(ws6, len(equiv) + 1, len(headers6))
auto_width(ws6, len(headers6))

# ═══════════════════════════════════════════════════════════════
# SHEET 7: Related Types
# ═══════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Related Types")
headers7 = ["Type", "Related Types (score 0.7)", "Shape Class", "Shape Cross-Compat (score 0.4)"]
for c, h in enumerate(headers7, 1):
    ws7.cell(row=1, column=c, value=h)

types = [
    ["Tree", "Mega Tree, Spiral Tree", "triangle", "custom, linear"],
    ["Mega Tree", "Tree, Spiral Tree", "triangle", "custom, linear"],
    ["Spiral Tree", "Tree, Mega Tree", "triangle", "custom, linear"],
    ["Arch", "Candy Cane, Line", "linear", "custom, triangle"],
    ["Candy Cane", "Arch, Line", "linear", "custom, triangle"],
    ["Spinner", "Wreath, Snowflake, Circle", "circular", "custom, point"],
    ["Wreath", "Spinner, Circle", "circular", "custom, point"],
    ["Circle", "Wreath, Spinner", "circular", "custom, point"],
    ["Star", "Snowflake, Circle", "point", "custom, circular"],
    ["Snowflake", "Star, Spinner, Wreath", "point", "custom, circular"],
    ["Line", "Roofline, Outline, Flood, Pole", "linear", "custom, triangle"],
    ["Roofline", "Line, Outline", "linear", "custom, triangle"],
    ["Outline", "Line, Roofline", "linear", "custom, triangle"],
    ["Icicles", "Line, Roofline", "linear", "custom, triangle"],
    ["Pole", "Line", "linear", "custom, triangle"],
    ["Flood", "Line", "point", "custom, circular"],
    ["Matrix", "Fence", "matrix", "custom"],
    ["Fence", "Matrix, Pixel Forest", "matrix", "custom"],
    ["Pixel Forest", "Matrix, Fence", "matrix", "custom"],
    ["Sign", "Matrix", "matrix", "custom"],
    ["Spider", "Custom", "custom", "all shapes at 0.4"],
    ["Ghost", "Custom, Pumpkin", "custom", "all shapes at 0.4"],
    ["Pumpkin", "Custom, Ghost", "custom", "all shapes at 0.4"],
    ["Tombstone", "Custom", "custom", "all shapes at 0.4"],
    ["Bat", "Custom", "custom", "all shapes at 0.4"],
    ["Present", "Custom", "custom", "all shapes at 0.4"],
    ["Singing Face", "Custom", "custom", "all shapes at 0.4"],
    ["Window", "(none)", "matrix", "custom"],
]

for r, row in enumerate(types, 2):
    for c, val in enumerate(row, 1):
        ws7.cell(row=r, column=c, value=val)

style_header(ws7, len(headers7))
style_body(ws7, len(types) + 1, len(headers7))
auto_width(ws7, len(headers7))

# ═══════════════════════════════════════════════════════════════
# SHEET 8: Vendor Pixel Hints
# ═══════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Vendor Pixel Hints")
headers8 = ["Pixel Count", "Product ID", "Vendor", "Effect"]
for c, h in enumerate(headers8, 1):
    ws8.cell(row=1, column=c, value=h)

hints = [
    [269, "ccc_spinner_18", "CCC", "Score ×1.15 when both models have this pixel count"],
    [451, "ccc_spinner_24", "CCC", "Score ×1.15"],
    [519, "ccc_spinner_25", "CCC", "Score ×1.15"],
    [596, "ccc_spinner_36", "CCC", "Score ×1.15"],
    [768, "ccc_spinner_36_v2", "CCC", "Score ×1.15"],
    [1046, "ccc_spinner_48", "CCC", "Score ×1.15"],
    [640, "ge_flake_640", "Gilbert Engineering", "Score ×1.15"],
    [960, "ge_fuzion", "Gilbert Engineering", "Score ×1.15"],
    [1200, "ge_rosa_grande", "Gilbert Engineering", "Score ×1.15"],
    [1529, "ge_overlord", "Gilbert Engineering", "Score ×1.15"],
    [1800, "ge_click_click_boom", "Gilbert Engineering", "Score ×1.15"],
    [800, "efl_showstopper", "EFL Designs", "Score ×1.15"],
    [400, "efl_babyflake", "EFL Designs", "Score ×1.15"],
    [1117, "boscoyo_mesmerizer", "Boscoyo", "Score ×1.15"],
    [720, "boscoyo_whimsical", "Boscoyo", "Score ×1.15"],
    [480, "holidaycoro_24_spinner", "Holiday Coro", "Score ×1.15"],
    [512, "p10_matrix_16x32", "Generic Panel", "Score ×1.15"],
    [1024, "p10_matrix_32x32", "Generic Panel", "Score ×1.15"],
    [2048, "p5_matrix_64x32", "Generic Panel", "Score ×1.15"],
    [4096, "p5_matrix_64x64", "Generic Panel", "Score ×1.15"],
    [50, "standard_small_prop", "Standard", "Score ×1.15"],
    [100, "standard_medium_prop", "Standard", "Score ×1.15"],
    [150, "standard_custom_prop", "Standard", "Score ×1.15"],
    [200, "standard_candy_cane", "Standard", "Score ×1.15"],
    [250, "standard_tree_250", "Standard", "Score ×1.15"],
    [300, "standard_arch_300", "Standard", "Score ×1.15"],
    [500, "standard_outline_500", "Standard", "Score ×1.15"],
]

for r, row in enumerate(hints, 2):
    for c, val in enumerate(row, 1):
        ws8.cell(row=r, column=c, value=val)

style_header(ws8, len(headers8))
style_body(ws8, len(hints) + 1, len(headers8))
auto_width(ws8, len(headers8))

# ═══════════════════════════════════════════════════════════════
# SHEET 9: Interchangeable Classes
# ═══════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Interchangeable Classes")
headers9 = ["Class", "Keywords", "Phase Used", "Description", "Example Cross-Match"]
for c, h in enumerate(headers9, 1):
    ws9.cell(row=1, column=c, value=h)

classes = [
    ["halloween_yard", "pumpkin, ghost, tombstone, skeleton, spider, bat, skull, witch, zombie, reaper, scarecrow, cauldron, coffin, black cat",
     "Phase 3 + 3b", "Halloween yard props — interchangeable when no exact match exists",
     "8 Pumpkins (source) → 7 Ghosts (user) when no pumpkins in user layout"],
    ["christmas_yard", "snowman, present, gift, candy cane, stocking, nutcracker, ornament",
     "Phase 3 + 3b", "Christmas yard props — interchangeable for coverage",
     "Snowman → Present when no snowman available"],
    ["mini_yard", "mini tree, ghost, black cat, mini pumpkin",
     "Phase 3 + 3b", "Small yard decorations — cross-holiday interchangeable",
     "Mini Trees (Xmas) → Ghosts (Halloween) for coverage"],
    ["tree", "tree, mega tree, megatree, spiral",
     "Phase 3 + 3b", "All tree types are interchangeable",
     "Mega Tree → Spiral Tree when no mega tree in user layout"],
    ["arch", "arch, archway, candy cane, cane",
     "Phase 3 + 3b", "Arch-family props",
     "Arches → Candy Canes when no arches available"],
    ["star_flake", "star, snowflake, flake",
     "Phase 3 + 3b", "Star and snowflake props",
     "Stars → Snowflakes"],
    ["wreath_spinner", "wreath, spinner, rosa, fuzion",
     "Phase 3 + 3b", "Circular spinning/wreath props",
     "Wreaths → Spinners"],
    ["structural_line", "eave, vertical, roofline, outline, horizontal",
     "Phase 3 + 3b", "Linear structural elements",
     "Eaves → Roofline"],
    ["vertical_structure", "pole, vertical, fence",
     "Phase 3 + 3b", "Vertical structural elements",
     "Poles → Verticals"],
]

for r, row in enumerate(classes, 2):
    for c, val in enumerate(row, 1):
        ws9.cell(row=r, column=c, value=val)

style_header(ws9, len(headers9))
style_body(ws9, len(classes) + 1, len(headers9))
auto_width(ws9, len(headers9))

# ═══════════════════════════════════════════════════════════════
# SHEET 10: Score Modifiers
# ═══════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("Score Modifiers")
headers10 = ["#", "Modifier Name", "Condition", "Effect", "Rationale"]
for c, h in enumerate(headers10, 1):
    ws10.cell(row=1, column=c, value=h)

modifiers = [
    [1, "Relaxed Pixel Scoring (V5.2)", "Non-group, name ≥0.85, type ≥0.7", "Floor pixel factor at 0.5",
     "When models clearly match by name and type, pixel count differences shouldn't drag them to low confidence"],
    [2, "House-Location Reweighting", "Either model has house location term (office, garage, door, etc.) and neither is a group",
     "Reweight: name 0.23, spatial 0.22, shape 0.13, type 0.25, pixels 0.10, structure 0.07 (name ↓, type ↑)",
     "Location-specific models should prioritize TYPE over NAME to avoid 'Garage Peak' matching 'Garage Door'"],
    [3, "Coordinate Tiebreaker", "Same base name, different indices, both non-group",
     "Reweight: name 0.18, spatial 0.42 (spatial becomes primary)",
     "When two Arch 3s exist, position should determine which maps to which Arch"],
    [4, "Eave/Vert Deprioritization", "Source is eave/vert individual and dest is NOT a house-line model",
     "Score × 0.4",
     "Eaves/verts are structural elements that shouldn't match decorative props"],
    [5, "Vendor Pixel Hint Boost", "Both non-group, same pixelCount, count is in VENDOR_PIXEL_HINTS",
     "Score × 1.15 (capped at 1.0)",
     "Models with known vendor pixel counts are likely the same physical product"],
    [6, "Submodel Group Semantic Match", "Both SUBMODEL_GROUPs with same semanticCategory",
     "Score = 0.75 + name*0.25 (capped at 1.0)",
     "Cross-vendor spinner groups with same semantic function (e.g., both 'florals') should match highly"],
    [7, "Submodel Group Category Mismatch", "Both SUBMODEL_GROUPs, both have categories but they differ",
     "Score = name * 0.5 (penalized)",
     "Spinner submodel groups with different functions (e.g., 'spokes' vs 'rings') should be penalized"],
    [8, "Surplus-to-Spatial Fallback (V5.2)", "Unmapped individual, shares base name with matched dest",
     "Fixed score = 0.65 (medium confidence)",
     "When source has more props than user, surplus maps to nearest same-type dest via many-to-one"],
]

for r, row in enumerate(modifiers, 2):
    for c, val in enumerate(row, 1):
        ws10.cell(row=r, column=c, value=val)

style_header(ws10, len(headers10))
style_body(ws10, len(modifiers) + 1, len(headers10))
auto_width(ws10, len(headers10))

# ═══════════════════════════════════════════════════════════════
# SHEET 11: Confidence Tiers
# ═══════════════════════════════════════════════════════════════
ws11 = wb.create_sheet("Confidence Tiers")
headers11 = ["Tier", "Score Range", "Auto-Accept?", "UI Treatment", "Description"]
for c, h in enumerate(headers11, 1):
    ws11.cell(row=1, column=c, value=h)

tiers = [
    ["high", "≥ 0.85", "Yes (Phase 1 AutoAccept)", "Green badge, pre-accepted",
     "Very confident match — name, type, position all align strongly"],
    ["medium", "≥ 0.60 and < 0.85", "No", "Yellow badge, needs user confirmation",
     "Reasonable match but one or more factors are weak. Common for cross-type matches, pixel drift, or spatial mismatch."],
    ["low", "≥ 0.40 and < 0.60", "No", "Red badge, shown as suggestion only",
     "Weak match — significant differences in name, type, or position. User should verify manually."],
    ["unmapped", "< 0.40", "No", "Not shown as match, appears in unmatched list",
     "No viable match found. Model needs manual drag-and-drop assignment."],
]

for r, row in enumerate(tiers, 2):
    for c, val in enumerate(row, 1):
        ws11.cell(row=r, column=c, value=val)
    # Color-code
    tier_cell = ws11.cell(row=r, column=1)
    if row[0] == "high":
        tier_cell.fill = HIGH_FILL
    elif row[0] == "medium":
        tier_cell.fill = MED_FILL
    elif row[0] == "low":
        tier_cell.fill = LOW_FILL

style_header(ws11, len(headers11))
style_body(ws11, len(tiers) + 1, len(headers11))
auto_width(ws11, len(headers11))

# ═══════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════
output_path = "/home/user/Lights-of-Elm-Ridge/ModIQ_V5.2_Matching_Rules.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Sheets: {', '.join(wb.sheetnames)}")
